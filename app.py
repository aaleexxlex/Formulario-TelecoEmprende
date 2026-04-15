from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    session
)
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
from time import time
import os
import re
import secrets
from html import escape

app = Flask(__name__)

# Usa una clave segura desde variable de entorno.
# Si no existe, genera una temporal para desarrollo.
app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))

# Configuración de cookies de sesión más segura
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = False  # Ponlo en True cuando uses HTTPS real

EXCEL_FILE = Path("registros_evento.xlsx")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")

# Rate limiting simple en memoria
request_log = {}

# Configuración básica anti abuso
MAX_REQUESTS_PER_MINUTE = 8
BLOCK_WINDOW_SECONDS = 60
MAX_NOMBRE_LEN = 60
MAX_APELLIDOS_LEN = 100
MAX_ESTUDIOS_LEN = 120
MAX_EMAIL_LEN = 120


def crear_excel_si_no_existe():
    if not EXCEL_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Inscripciones"
        ws.append([
            "Nombre",
            "Apellidos",
            "Estudios / Procedencia",
            "Correo electrónico",
            "Acepta privacidad",
            "IP registro",
            "Fecha de registro"
        ])
        ajustar_columnas(ws)
        wb.save(EXCEL_FILE)


def ajustar_columnas(ws):
    anchos = {
        "A": 20,
        "B": 25,
        "C": 35,
        "D": 32,
        "E": 18,
        "F": 18,
        "G": 22,
    }
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho


def limpiar_texto(texto: str) -> str:
    texto = " ".join(texto.strip().split())
    return escape(texto)


def email_valido(email: str) -> bool:
    patron = r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
    return re.match(patron, email) is not None


def longitud_valida(nombre: str, apellidos: str, estudios: str, email: str) -> bool:
    return (
        len(nombre) <= MAX_NOMBRE_LEN and
        len(apellidos) <= MAX_APELLIDOS_LEN and
        len(estudios) <= MAX_ESTUDIOS_LEN and
        len(email) <= MAX_EMAIL_LEN
    )


def obtener_ip_real() -> str:
    # Si detrás hay proxy, intenta usar X-Forwarded-For
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.remote_addr or "desconocida"


def demasiadas_peticiones(ip: str) -> bool:
    ahora = time()

    if ip not in request_log:
        request_log[ip] = []

    # Conserva solo peticiones recientes
    request_log[ip] = [t for t in request_log[ip] if ahora - t < BLOCK_WINDOW_SECONDS]

    if len(request_log[ip]) >= MAX_REQUESTS_PER_MINUTE:
        return True

    request_log[ip].append(ahora)
    return False


def email_ya_registrado(email: str) -> bool:
    if not EXCEL_FILE.exists():
        return False

    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb["Inscripciones"]

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[3] and str(fila[3]).strip().lower() == email.strip().lower():
            return True

    return False


def guardar_registro(nombre: str, apellidos: str, estudios: str, email: str, acepta_privacidad: str, ip: str):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Inscripciones"]

    fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([
        nombre,
        apellidos,
        estudios,
        email.lower(),
        acepta_privacidad,
        ip,
        fecha_registro
    ])

    ajustar_columnas(ws)
    wb.save(EXCEL_FILE)


def obtener_registros():
    if not EXCEL_FILE.exists():
        return []

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Inscripciones"]

    registros = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if any(fila):
            registros.append({
                "nombre": fila[0] or "",
                "apellidos": fila[1] or "",
                "estudios": fila[2] or "",
                "email": fila[3] or "",
                "privacidad": fila[4] or "",
                "ip": fila[5] or "",
                "fecha": fila[6] or "",
            })

    return registros


@app.after_request
def aplicar_headers_seguridad(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    response.headers["Cache-Control"] = "no-store"
    # CSP sencilla. Si luego añades scripts inline o más servicios externos, habrá que ajustarla.
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "img-src 'self' data:; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src https://fonts.gstatic.com; "
        "script-src 'self'; "
        "connect-src 'self'; "
        "frame-ancestors 'none';"
    )
    return response


@app.route("/", methods=["GET", "POST"])
def index():
    crear_excel_si_no_existe()

    if request.method == "POST":
        ip = obtener_ip_real()

        # Rate limit por IP
        if demasiadas_peticiones(ip):
            flash("Demasiadas solicitudes desde esta red. Inténtalo en un minuto.", "error")
            return redirect(url_for("index"))

        # Honeypot anti-bot
        telefono_oculto = request.form.get("telefono_oculto", "").strip()
        if telefono_oculto:
            # No damos feedback para no ayudar a bots
            return redirect(url_for("index"))

        nombre = limpiar_texto(request.form.get("nombre", ""))
        apellidos = limpiar_texto(request.form.get("apellidos", ""))
        estudios = limpiar_texto(request.form.get("estudios", ""))
        email = limpiar_texto(request.form.get("email", "")).lower()
        acepta_privacidad = request.form.get("privacidad")

        if not nombre or not apellidos or not estudios or not email:
            flash("Completa todos los campos obligatorios.", "error")
            return redirect(url_for("index"))

        if not longitud_valida(nombre, apellidos, estudios, email):
            flash("Alguno de los campos supera la longitud permitida.", "error")
            return redirect(url_for("index"))

        if not email_valido(email):
            flash("Introduce un correo electrónico válido.", "error")
            return redirect(url_for("index"))

        if not acepta_privacidad:
            flash("Debes aceptar la política de privacidad para registrarte.", "error")
            return redirect(url_for("index"))

        if email_ya_registrado(email):
            flash("Ese correo ya está registrado.", "error")
            return redirect(url_for("index"))

        guardar_registro(
            nombre=nombre,
            apellidos=apellidos,
            estudios=estudios,
            email=email,
            acepta_privacidad="Sí",
            ip=ip
        )

        flash("Registro completado. Te contactaremos pronto con la información del evento.", "success")
        return redirect(url_for("index"))

    return render_template("index.html")


@app.route("/admin", methods=["GET", "POST"])
def admin():
    crear_excel_si_no_existe()

    if not ADMIN_PASSWORD:
        flash("El panel admin no está configurado correctamente.", "error")
        return render_template("admin.html", autenticado=False, registros=[])

    if request.method == "POST":
        ip = obtener_ip_real()

        if demasiadas_peticiones(ip):
            flash("Demasiados intentos. Espera un minuto.", "error")
            return redirect(url_for("admin"))

        password = request.form.get("password", "")
        if password == ADMIN_PASSWORD:
            session.clear()
            session["admin_auth"] = True
            return redirect(url_for("admin"))

        flash("Contraseña incorrecta.", "error")
        return redirect(url_for("admin"))

    autenticado = session.get("admin_auth", False)

    if not autenticado:
        return render_template("admin.html", autenticado=False, registros=[])

    registros = obtener_registros()
    total_registros = len(registros)

    return render_template(
        "admin.html",
        autenticado=True,
        registros=registros,
        total_registros=total_registros
    )


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    flash("Sesión cerrada correctamente.", "success")
    return redirect(url_for("admin"))


@app.route("/admin/descargar")
def descargar_excel():
    autenticado = session.get("admin_auth", False)
    if not autenticado:
        flash("No autorizado.", "error")
        return redirect(url_for("admin"))

    if not EXCEL_FILE.exists():
        flash("No existe ningún archivo de registros todavía.", "error")
        return redirect(url_for("admin"))

    return send_file(
        EXCEL_FILE,
        as_attachment=True,
        download_name="registros_telecoemprende.xlsx"
    )


if __name__ == "__main__":
    crear_excel_si_no_existe()
    app.run(host="0.0.0.0", port=5000, debug=True)